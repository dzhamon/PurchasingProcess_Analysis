import os
import pandas as pd
from sklearn.cluster import KMeans
from sklearn.preprocessing import StandardScaler
from utils.config import BASE_DIR
from PyQt5.QtWidgets import QMessageBox

from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
from sklearn.decomposition import PCA
import matplotlib.pyplot as plt


def unique_discip_actor_lots(df):
	from utils.functions import CurrencyConverter
	
	converter = CurrencyConverter()
	df = converter.convert_column(
		df,
		amount_column='total_price',
		currency_column='currency',
		result_column='total_price_eur'
	).copy()
	
	unique_disciplines = df['discipline'].unique()
	lots_per_actor = {}
	
	for discipline in unique_disciplines:
		df_filtered = df[df['discipline'] == discipline].copy()
		
		if not df_filtered.empty:
			# Группируем по actor_name и winner_name, считаем лоты и сумму
			grouped = df_filtered.groupby(['actor_name', 'winner_name', 'lot_number']).agg(
				lot_count=('lot_number', 'count'),
				total_eur=('total_price_eur', 'sum')
			).reset_index()
			
			# Фильтруем записи с lot_count > 0
			grouped = grouped[grouped['lot_count'] > 0]
			
			# Сохраняем в виде словаря
			discipline_data = {}
			for _, row in grouped.iterrows():
				actor = row['actor_name']
				winner = row['winner_name']
				lotnum = row['lot_number']
				
				if actor not in discipline_data:
					discipline_data[actor] = {}
				
				discipline_data[actor][winner] = {
					'lot_count': row['lot_count'],
					'lotnum': row['lot_number'],
					'total_eur': row['total_eur']
				}
			
			lots_per_actor[discipline] = discipline_data
	
	return lots_per_actor


def analyze_suppliers(parent_widget, lots_per_actor):
	# 1. Собираем всех уникальных поставщиков по всем дисциплинам
	all_suppliers = set()
	lot_details = []
	for discipline, actors in lots_per_actor.items():
		for actor, suppliers in actors.items():
			all_suppliers.update(suppliers.keys())
			for supplier, info in suppliers.items():
				lot_details.append({
					'discipline': discipline,
					'actor': actor,
					'supplier': supplier,
					'lotnum': info['lotnum'],
					'lot_count': info['lot_count'],
					'total_eur': info['total_eur']
				})
	all_suppliers = list(all_suppliers)  # здесь есть lotnums?
	lots_df = pd.DataFrame(lot_details)  # Детальная таблица лотов
	
	# 2. Создаём агрегированный DataFrame для кластеризации
	agg_data = []
	for discipline, actors in lots_per_actor.items():
		for actor, suppliers in actors.items():
			row = {
				'discipline': discipline,
				'actor': actor,
				'total_eur': sum(info['total_eur'] for info in suppliers.values()),
				'total_lots': sum(info['lot_count'] for info in suppliers.values()),
				'unique_lotnums': len({info['lotnum'] for info in suppliers.values()}),
				'suppliers_count': len(suppliers)
			}
			# Добавляем поставщиков как бинарные признаки
			for supplier in all_suppliers:
				row[f'has_{supplier}'] = 1 if supplier in suppliers else 0
			
			agg_data.append(row)
	
	df = pd.DataFrame(agg_data)
	
	# 3. Нормализация данных для кластеризации
	numeric_cols = ['total_eur', 'total_lots', 'suppliers_count'] + [f'has_{s}' for s in
	                                                                 all_suppliers]  # здесь внимательно с lotnums
	scaler = StandardScaler()
	df_scaled = df[numeric_cols].copy()
	df_scaled[numeric_cols] = scaler.fit_transform(df[numeric_cols])
	
	# 4. Кластеризация (пример с K-Means)
	kmeans = KMeans(n_clusters=4, random_state=42)
	df['cluster'] = kmeans.fit_predict(df_scaled)
	
	# 5. Анализ результатов
	print("\nСтатистика по кластерам:")
	cluster_stats = df.groupby('cluster').agg({
		'actor': 'count',
		'total_eur': ['mean', 'median'],
		'total_lots': 'mean',
	})
	
	cols = df.columns.tolist()
	cols.remove('cluster')
	cols.insert(cols.index('total_lots') + 1, 'cluster')
	clustered_data = df[cols]
	
	print('clustered_data', clustered_data.columns)  # Агрегированные данные с кластерами
	print('lot_details', lots_df.columns)  # Детализация по лотам
	print('supplier_matrix', df_scaled.columns)  # Матрица для анализа
	
	from excel_tables.ExcelExporter import ExcelExporter
	
	dataframes = {
		'clustered': clustered_data,
		'lot_details': lots_df,
		'supplier_matrix': df_scaled
	}
	
	# Задаем путь к директории выходного файла
	OUT_DIR = os.path.join(BASE_DIR, 'cluster_results')
	os.makedirs(OUT_DIR, exist_ok=True)
	
	file_name = 'clustering_results.xlsx'
	file_path = os.path.join(OUT_DIR, file_name)
	
	with ExcelExporter(file_path) as exporter:
		exporter.save_multiple_dataframes(
			dataframes,
			sheet_names={
				'clustered': 'Кластеризованные данные',
				'lot_details': 'Детали лотов',
				'supplier_matrix': 'Матрица поставщиков'
			},
			index=False, float_format='%.2f'
		)
		# Дополнительное форматирование через openpyxl
		wb = exporter.writer.book
		num_format = '#,##0.00'
		
		for sheet_name in wb.sheetnames:
			ws = wb[sheet_name]
			
			# 1. Добавляем заголовок таблицы
			ws.insert_rows(1)  # Вставляем новую первую строку
			ws['A1'] = f"Таблица: {sheet_name}"
			ws['A1'].font = Font(bold=True, size=12)
			
			# 2. Применяем числовой формат ко всем числовым ячейкам
			for row in ws.iter_rows(min_row=3):  # Данные начинаются с 3 строки
				for cell in row:
					if isinstance(cell.value, (float)):
						cell.number_format = num_format
	
	QMessageBox.information(parent_widget, "Результат",
	                        f"Классификация исполнителей по поставщикам завершена. Файл сохранен в папке "
	                        f"{OUT_DIR} ")
	return
