import pandas as pd
from collections import defaultdict
import os
from utils.config import BASE_DIR
from openpyxl.styles import Font, Alignment, NamedStyle

def calculate_statistics(warehouse_df):
	print("Входим в метод расчета статистик по Складам")
	# Расчет количество складов за выделенный период
	# warehouses_count = warehouse_df['warehouse'].value_counts() # подсчет количества встречаемости в столбце его значений
	warehouses_unique_count = len(warehouse_df['warehouse'].unique())
	print("Количество складов = ", warehouses_unique_count)
	
	# Общее количество товарной номенклатуры на складах на последнюю дату
	max_date = warehouse_df["date_column"].max()
	print("Последняя дата --> ", max_date)
	unique_nomenclature_count = warehouse_df.loc[warehouse_df["date_column"] == max_date, "nomenclature"].nunique()
	print(f"Общее количество товарной номенклатуры на дату {max_date} = {unique_nomenclature_count}")
	
	# Считаем товарные запасы по отдельным складам общей суммой в размере валют
	# Фильтруем данные по максимальной дате
	filtered_df = warehouse_df[warehouse_df['date_column'] == max_date]
	
	# Группируем по складу и валюте, затем суммируем total_price_currency
	result_df = filtered_df.groupby(['warehouse', 'currency'])['total_price_currency'].sum().reset_index()
	
	# Выводим полученный результат в файл excel
	from excel_tables.ExcelExporter import ExcelExporter
	
	# Задаем путь к директории выходного файла
	OUT_DIR = os.path.join(BASE_DIR, 'warehouse_totals')
	os.makedirs(OUT_DIR, exist_ok=True)
	file_name = 'warehouse_totals_formatted.xlsx'
	file_path = os.path.join(OUT_DIR, file_name)
	
	with ExcelExporter(file_path) as exporter:
		exporter.save_dataframe(
			df=result_df, # пересылаемый датафрейм
			sheet_name="Отчет по остаткам на складах",
			index=False,
			float_format="%.2f"
		)
		# дополнительное форматирование через openpyxl
		wb = exporter.writer.book
		num_format = '#,##0.00'
		
		ws = wb['sheet_name']
		ws.insert_rows(1)
		ws['A1'] = 'Остатки на складе в валютах контракта'
		ws['A1'].font = Font(name='Arial Narrow', size=14, bold=True)
		ws['A1'].alignment = Alignment(horizontal='center')
		
	return